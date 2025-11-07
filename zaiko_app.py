# zaikokanri1_app.py
import streamlit as st
import pandas as pd
from datetime import datetime
import os
import tempfile
import shutil
from openpyxl import load_workbook, Workbook

# ---------- 設定 ----------
BASE_DIR = os.getcwd()  # Streamlit Cloud ではリポジトリルートに置く想定

# あなたの既存ファイル名（repo に置くか、ローカル運用ならフルパスに）
EXCEL_FILE = os.path.join(BASE_DIR, "原料在庫表.xlsx")   # .xlsm を .xlsx にしておくと安全
MATERIAL_MASTER = os.path.join(BASE_DIR, "material_master.xlsx")
STAFF_MASTER = os.path.join(BASE_DIR, "staff_master.xlsx")
LOG_FILE = os.path.join(BASE_DIR, "inventory_log.xlsx")

# ----- シンプル認証（コード内パスワード） -----
PASSWORD = "takaki2226"

def check_password():
    if "password_ok" not in st.session_state:
        st.session_state.password_ok = False

    if not st.session_state.password_ok:
        st.header("🔐 ログイン")
        st.text_input("パスワードを入力してください", type="password", key="password_input")
        if st.button("ログイン"):
            if st.session_state.get("password_input") == PASSWORD:
                st.session_state.password_ok = True
                st.rerun()   # ← 修正ポイント
            else:
                st.error("パスワードが違います")
        st.stop()

check_password()

# ---------- ヘルパー関数 ----------
def read_master_from_excel(excel_path, keyword):
    """EXCEL_FILE の中の列名に keyword を含む列を探してリストを返す"""
    if not os.path.exists(excel_path):
        return []
    try:
        xls = pd.read_excel(excel_path, sheet_name=None)
        for name, df in xls.items():
            cols = [c for c in df.columns if isinstance(c, str) and keyword in c]
            if cols:
                return df[cols[0]].dropna().astype(str).unique().tolist()
        return []
    except Exception as e:
        st.error(f"{os.path.basename(excel_path)} 読込エラー: {e}")
        return []

def safe_read_list(path, keyword, fallback=None):
    """material_master.xlsx などを優先読み、それが無ければ EXCEL_FILE を探す"""
    if os.path.exists(path):
        try:
            df = pd.read_excel(path)
            col = [c for c in df.columns if isinstance(c, str) and keyword in c]
            if col:
                return df[col[0]].dropna().astype(str).unique().tolist()
            if df.shape[1] >= 1:
                return df.iloc[:,0].dropna().astype(str).unique().tolist()
        except Exception as e:
            st.warning(f"{os.path.basename(path)} 読込時警告: {e}")
    if fallback and os.path.exists(fallback):
        return read_master_from_excel(fallback, keyword)
    return []

def ensure_logfile():
    if not os.path.exists(LOG_FILE):
        df_init = pd.DataFrame(columns=[
            "日時", "棚", "列", "段", "サブ", "材料名",
            "操作", "数量(kg)", "残数(kg)", "作業者", "メモ"
        ])
        df_init.to_excel(LOG_FILE, index=False)

def sync_history_to_excel():
    """
    VBA の '履歴追加' と同等の処理を行う。
    inventory_log.xlsx の内容を読み込み、原料在庫表.xlsx の '履歴' シートを全置換する。
    """
    try:
        # inventory_log.xlsx を読む
        if not os.path.exists(LOG_FILE):
            st.warning("履歴ファイルが見つかりません（inventory_log.xlsx）。履歴反映をスキップします。")
            return

        df_history = pd.read_excel(LOG_FILE)

        # Excel ファイルが存在するか確認して workbook を作成／読み込み
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            # 履歴シートがある場合は中身を削除、無ければ作成
            if "履歴" in wb.sheetnames:
                ws = wb["履歴"]
                # 全行削除（ヘッダ含む）
                if ws.max_row > 0:
                    ws.delete_rows(1, ws.max_row)
            else:
                ws = wb.create_sheet("履歴")
        else:
            # 新規ブック作成して履歴シートを作る
            wb = Workbook()
            ws = wb.active
            ws.title = "履歴"

        # DataFrame のヘッダーを書き込み
        if df_history.shape[0] == 0:
            # 空データならヘッダだけ書く（列名がある場合）
            headers = list(df_history.columns) if df_history.shape[1] > 0 else ["日時","棚","列","段","サブ","材料名","操作","数量(kg)","残数(kg)","作業者","メモ"]
            ws.append(headers)
        else:
            headers = list(df_history.columns)
            ws.append(headers)
            for row in df_history.itertuples(index=False, name=None):
                # openpyxl は None を空セルとして扱うのでそのまま append で良い
                ws.append(list(row))

        # 保存（上書き）
        wb.save(EXCEL_FILE)
        st.info("📘 原料在庫表の『履歴』シートを更新しました（Python版）")
    except Exception as e:
        st.warning(f"⚠️ 履歴反映エラー: {e}")

# ---------- マスター読込 ----------
materials = safe_read_list(MATERIAL_MASTER, "原料", fallback=EXCEL_FILE)
staffs = safe_read_list(STAFF_MASTER, "作業者", fallback=EXCEL_FILE)

if not materials:
    materials = ["材料A", "材料B", "材料C"]
if not staffs:
    staffs = ["作業者A", "作業者B"]

ensure_logfile()

# ---------- UI ----------
st.title("📦 棚在庫管理（Web版）")

st.sidebar.header("操作")
mode = st.sidebar.selectbox("モードを選択", ["入出庫", "在庫一覧", "履歴"])

if mode == "入出庫":
    st.header("⚙️ 入出庫登録")
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        shelf = st.selectbox("棚", [1,2,3,4], index=0)
    with c2:
        row = st.selectbox("列", list(range(1,21)))
    with c3:
        level = st.selectbox("段", list(range(1,6)))
    with c4:
        sub = st.selectbox("サブ", ["", "1", "2"], index=0)

    # 現在の在庫表示
    try:
        df_log = pd.read_excel(LOG_FILE)
        df_log["サブ"] = pd.to_numeric(df_log["サブ"], errors="coerce").fillna(0).astype(int)
        df_log["棚"] = pd.to_numeric(df_log["棚"], errors="coerce").fillna(0).astype(int)
        df_log["列"] = pd.to_numeric(df_log["列"], errors="coerce").fillna(0).astype(int)
        df_log["段"] = pd.to_numeric(df_log["段"], errors="coerce").fillna(0).astype(int)
        sub_val = int(sub) if str(sub).isdigit() else 0

        df_loc = df_log[
            (df_log["棚"] == shelf) &
            (df_log["列"] == row) &
            (df_log["段"] == level) &
            (df_log["サブ"] == sub_val)
        ]
        if not df_loc.empty:
            last = df_loc.iloc[-1]
            cur_material = last["材料名"]
            cur_stock = float(last["残数(kg)"])
            st.info(f"🧾 現在の在庫：{cur_material}（{cur_stock} kg）")
        else:
            cur_material, cur_stock = None, 0
            st.warning("📭 この棚は空です。")
    except Exception as e:
        st.error(f"在庫情報取得エラー: {e}")
        cur_material, cur_stock = None, 0
        sub_val = int(sub) if str(sub).isdigit() else 0

    st.subheader("入出庫情報入力")
    c5, c6, c7, c8 = st.columns(4)
    with c5:
        operation = st.radio("操作", ["入庫","出庫"], horizontal=True)
    with c6:
        material = st.selectbox("材料名", materials)
    with c7:
        qty = st.number_input("数量 (kg)", min_value=1, max_value=999999, step=1, value=1)
    with c8:
        staff = st.selectbox("作業者名", staffs)

    if st.button("登録する"):
        qty_signed = qty if operation == "入庫" else -qty
        try:
            df_old = pd.read_excel(LOG_FILE)
            df_old["サブ"] = pd.to_numeric(df_old["サブ"], errors="coerce").fillna(0).astype(int)
            df_loc = df_old[
                (df_old["棚"] == shelf) &
                (df_old["列"] == row) &
                (df_old["段"] == level) &
                (df_old["サブ"] == sub_val)
            ]
            cur_stock = df_loc["残数(kg)"].iloc[-1] if not df_loc.empty else 0
            cur_material = df_loc.iloc[-1]["材料名"] if not df_loc.empty else None

            if operation == "出庫" and cur_stock <= 0:
                st.error("❌ 空の棚からは出庫できません。")
            elif cur_material and cur_material != material:
                st.error(f"❌ この棚には別の材料「{cur_material}」が登録されています。")
            else:
                new_stock = cur_stock + qty_signed
                if new_stock < 0:
                    st.error("⚠️ 出庫数量が在庫を超えています。")
                else:
                    new_entry = {
                        "日時": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "棚": shelf, "列": row, "段": level, "サブ": sub_val,
                        "材料名": material, "操作": operation,
                        "数量(kg)": qty_signed, "残数(kg)": new_stock,
                        "作業者": staff, "メモ": ""
                    }
                    df_updated = pd.concat([df_old, pd.DataFrame([new_entry])], ignore_index=True)

                    # 安全に一時ファイル経由で上書き保存
                    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                        tmp_path = tmp.name
                        df_updated.to_excel(tmp_path, index=False)
                    shutil.move(tmp_path, LOG_FILE)

                    st.success(f"✅ 登録完了！（残数：{new_stock} kg）")

                    # --- ここで VBA 履歴追加 と同等の処理を行う ---
                    sync_history_to_excel()

        except Exception as e:
            st.error(f"履歴保存に失敗しました: {e}")

elif mode == "在庫一覧":
    st.header("📋 在庫一覧")
    try:
        df_log = pd.read_excel(LOG_FILE)
        df_log["サブ"] = pd.to_numeric(df_log["サブ"], errors="coerce").fillna(0).astype(int)
        grouped = df_log.groupby(["棚","列","段","サブ","材料名"], dropna=False)["残数(kg)"].last().reset_index()
        st.dataframe(grouped.sort_values(["棚","列","段","サブ"]).reset_index(drop=True))
        csv = grouped.to_csv(index=False).encode("utf-8-sig")
        st.download_button("在庫一覧をCSVでダウンロード", data=csv, file_name="zaiko_current.csv", mime="text/csv")
    except Exception as e:
        st.error(f"在庫一覧取得エラー: {e}")

else:  # 履歴
    st.header("📜 履歴（入出庫ログ）")
    try:
        df_log = pd.read_excel(LOG_FILE)
        df_log["日時"] = pd.to_datetime(df_log["日時"], errors="coerce")
        col1, col2, col3 = st.columns(3)
        with col1:
            from_date = st.date_input("期間開始", value=pd.to_datetime("2000-01-01"))
        with col2:
            to_date = st.date_input("期間終了", value=pd.to_datetime("2100-01-01"))
        with col3:
            name_filter = st.text_input("材料名で絞り込み（部分一致）", value="")

        mask = (df_log["日時"].dt.date >= from_date) & (df_log["日時"].dt.date <= to_date)
        if name_filter:
            mask &= df_log["材料名"].astype(str).str.contains(name_filter)
        display = df_log[mask].sort_values("日時", ascending=False).reset_index(drop=True)
        st.dataframe(display)
        csv = display.to_csv(index=False).encode("utf-8-sig")
        st.download_button("履歴をCSVでダウンロード", data=csv, file_name="inventory_history.csv", mime="text/csv")
    except Exception as e:
        st.error(f"履歴取得エラー: {e}")
